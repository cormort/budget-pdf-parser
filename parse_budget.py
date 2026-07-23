"""
Budget-PDF parser, mirroring the logic proposed for index.html (v36).

Pipeline:
  1. Extract text from the PDF (ordered by y-position of each text chunk, like pdf.js).
  2. Apply three cleanup passes identical to the browser buttons:
       a) break lines by headings (壹/一/account-name-dash/N.)
       b) insert newline + 2-space indent before (N)/(一)
       c) collapse intra-line whitespace, preserve leading indent
  3. Parse each cleaned line into a structured row.
  4. Write rows to xlsx.
"""

import re
import sys
from functools import lru_cache
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ----- Account chart (identical to the one embedded in index.html) -----
ACCOUNT_CHART_TEXT = """\
1\t用人費用
11\t正式員額薪資
111\t管理會委員報酬
112\t顧問人員報酬
113\t職員薪金
114\t工員工資
115\t警餉
12\t聘僱及兼職人員薪資
121\t聘用人員薪金
122\t約僱職員薪金
123\t約僱工員薪資
124\t兼職人員酬金
13\t加(夜)班費
131\t延長工時加班費
132\t夜班費
133\t誤餐費
134\t未休假加班費
14\t津貼
141\t水電津貼
142\t領班津貼
143\t僻地津貼
144\t工地津貼
14Y\t其他津貼
15\t獎金
151\t考績獎金
152\t年終獎金
15Y\t其他獎金
16\t退休及卹償金
161\t職員退休及離職金
162\t工員退休及離職金
163\t卹償金
17\t資遣費
171\t職員資遣費
172\t工員資遣費
18\t福利費
181\t分擔員工保險費
182\t分擔退休人員及其配偶暨員工眷屬保險費
183\t傷病醫藥費
184\t提撥福利金
185\t員工通勤交通費
186\t分擔輔助建屋貸款利息
18Y\t其他福利費
19\t提繳費
191\t提繳工資墊償費用
2\t服務費用
21\t水電費
211\t動力費
212\t工作場所電費
213\t宿舍電費
214\t工作場所水費
215\t宿舍水費
216\t氣體費
21Y\t其他場所水電費
22\t郵電費
221\t郵費
222\t電話費
223\t電報費
224\t數據通信費
23\t旅運費
231\t國內旅費
232\t國外旅費
233\t大陸地區旅費
234\t專力費
235\t貨物運費
236\t裝卸費
237\t港埠費
23Y\t其他旅運費
24\t印刷裝訂及公告費
241\t印刷及裝訂費
242\t廣告費
243\t公告費
244\t樣品贈送
245\t業務宣導費
25\t修理保養及保固費
251\t土地改良物修護費
252\t一般房屋修護費
253\t宿舍修護費
254\t其他建築修護費
255\t機械及設備修護費
256\t交通及運輸設備修護費
257\t雜項設備修護費
258\t其他資產修護費
25Y\t保固費
26\t保險費
261\t一般房屋保險費
262\t宿舍保險費
263\t機械及設備保險費
264\t交通及運輸設備保險費
265\t雜項設備保險費
266\t其他資產保險費
267\t現金、存款及貨物保險費
268\t責任保險費
26Y\t其他保險費
27\t一般服務費
271\t棧儲費
272\t包裝費
273\t公證費
274\t報關費
275\t理貨費
276\t佣金、匯費、經理費及手續費
277\t代理(辦)費
278\t加工費
279\t外包費
27A\t義(志)工服務費
27D\t計時與計件人員酬金
27E\t替代役待遇及給與
27F\t體育活動費
27G\t核能除役服務費
28\t專業服務費
281\t技術合作費及權利金
282\t專技人員酬金
283\t法律事務費
284\t工程及管理諮詢服務費
285\t講課鐘點、稿費、出席審查及查詢費
286\t委託調查研究費
287\t委託檢驗(定)試驗認證費
288\t委託考選訓練費
289\t試務甄選費
28A\t電腦軟體服務費
28Y\t其他專業服務費
29\t公關慰勞費
291\t公共關係費
292\t員工慰勞費
2A\t媒體政策及業務宣導費
2A1\t媒體政策及業務宣導費
2B\t推展費
2B1\t推展費
3\t材料及用品費
31\t使用材料費
311\t物料
312\t燃料
313\t油脂
314\t建築材料
315\t設備零件
32\t用品消耗
321\t辦公(事務)用品
322\t報章雜誌
323\t農業與園藝用品及環境美化費
324\t化學藥劑與實驗用品
325\t服裝
326\t食品
327\t飼料
328\t醫療用品(非醫療院所使用)
32Y\t其他用品消耗
33\t商品
331\t商品
4\t租金、償債、利息及相關手續費
41\t地租及水租
411\t一般土地租金
412\t宿舍基地租金
413\t場地租金
42\t房租
421\t一般房屋租金
422\t宿舍租金
43\t機器租金
431\t電腦租金及使用費
432\t機械及設備租金
44\t交通及運輸設備租金
441\t船租
442\t車租
443\t電信設備租金
444\t碼頭設備租金
445\t航空器租金
446\t貨櫃及車架租金
45\t雜項設備租金
451\t雜項設備租金
46\t償債、利息及相關手續費
461\t債務還本
462\t債券發行成本
463\t債務利息
464\t債券手續費
46Y\t其他利息
5\t購建固定資產、無形資產、非理財目的之長期投資及營舍與設施工程支出
51\t購建固定資產
511\t購置土地
512\t興建土地改良物
513\t擴充改良房屋建築及設備
514\t購置機械及設備
515\t購置交通及運輸設備
516\t購置雜項設備
517\t租賃資產租金
518\t租賃權益改良支出
519\t購置收藏品及傳承資產
52\t購置無形資產
521\t購置電腦軟體
522\t購置權利
53\t非理財目的之長期投資
531\t非理財目的之長期證券
532\t其他非理財目的之長期投資
54\t遞延支出
541\t遞延修繕房屋建築支出
542\t其他遞延支出
55\t營舍與設施工程支出
551\t營舍與設施工程支出
6\t稅捐及規費(強制費)
61\t土地稅
611\t土地增值稅
612\t一般土地地價稅
613\t宿舍基地地價稅
61Y\t其他土地地價稅
62\t契稅
621\t契稅
63\t房屋稅
631\t一般房屋稅
632\t宿舍房屋稅
63Y\t其他房屋稅
64\t消費與行為稅
641\t關稅
642\t貨物稅
643\t證券交易稅
644\t營業稅
645\t印花稅
646\t使用牌照稅
65\t特別稅課
651\t特別稅課
66\t規費
661\t行政規費與強制費
662\t事業規費
663\t汽車燃料使用費
664\t商港服務費
665\t未足額進用身障人員差額補助費
666\t碳費
66Y\t其他規費
7\t會費、捐助、補助、分攤、照護、救濟與交流活動費
71\t會費
711\t國際組織會費
712\t學術團體會費
713\t職業團體會費
72\t捐助、補助與獎助
721\t補(協)助政府機關(構)
722\t捐助國內團體
723\t捐助私校
724\t捐助個人
725\t對外國之捐助
726\t獎助學員生給與
72Y\t其他捐助、補助與獎助
73\t分攤
731\t分擔污染防制費
732\t分擔大樓管理費
733\t分擔礦場保安費
734\t分擔職業訓練費
73Y\t分擔其他費用
74\t補貼、獎勵、慰問、照護與救濟
741\t補貼就業訓練津貼與貸(存)款利息
742\t補貼收容人膳宿費、保險及資遣費
743\t獎勵費用
744\t慰問、照護及濟助金
745\t醫療衛生受害救濟給付
74Y\t其他補貼、獎勵、慰問、照護與救濟
75\t競賽及交流活動費
751\t技能競賽
752\t交流活動費
8\t短絀、賠償給付與支應退場支出
81\t各項短絀
811\t磅(現金分)差
812\t呆帳及保證短絀
813\t運輸及搬運短絀
814\t停工短絀
815\t損壞工作
816\t災害短絀
817\t資產短絀
818\t兌換短絀
819\t投資短絀
81Y\t其他短絀
82\t賠償給付
821\t一般賠償
822\t旅運賠償
823\t公害賠償
83\t支應退場支出
831\t支應退場支出
84\t緊急應變支出
841\t緊急應變支出
85\t運動彩券發行賠損支出
851\t運動彩券發行賠損支出
9\t其他
91\t其他支出
912\t取得經營不善金融機構資產
91Y\t其他
"""


# ----- Helpers -----

@lru_cache(maxsize=None)
def normalize(s: str) -> str:
    """Strip chars that commonly differ between PDF text and the chart.

    - Full/half-width parenthesized content is removed (e.g. 補(協)助... -> 補助...)
    - Whitespace / CJK delimiter 、 / 及 / full-width dash 、 hyphens are removed.
    """
    s = re.sub(r"（[^）]*）", "", s)
    s = re.sub(r"\([^)]*\)", "", s)
    s = re.sub(r"[\s、及－\-]", "", s)
    return s


DROPS_RE = re.compile(r"[\s、及－\-]")


def _walk_match(line: str, norm_name: str) -> tuple[str, str] | None:
    """Walk the original string `line` consuming `len(norm_name)` normalized chars.
    Returns (code, name, rest) or None."""
    pos = 0
    consumed = 0
    target = len(norm_name)
    depth = 0
    while pos < len(line) and consumed < target:
        c = line[pos]
        if c in "（(":
            depth += 1
            pos += 1
            continue
        if c in "）)":
            if depth > 0:
                depth -= 1
            pos += 1
            continue
        if depth > 0:
            pos += 1
            continue
        if DROPS_RE.match(c):
            pos += 1
            continue
        consumed += 1
        pos += 1
    return line[pos:] if consumed == target else None


def build_chart():
    code_to_name = {}
    items: list[dict] = []
    for line in ACCOUNT_CHART_TEXT.strip().split("\n"):
        parts = line.split("\t")
        if len(parts) >= 2:
            code, name = parts[0].strip(), parts[1].strip()
            code_to_name[code] = name
            norm = normalize(name)
            if norm:
                items.append({"code": code, "name": name, "norm": norm})
    items.sort(key=lambda x: len(x["norm"]), reverse=True)

    # Pre-index by code length for O(1) level filtering
    by_len: dict[int, list[dict]] = {1: [], 2: [], 3: []}
    for item in items:
        clen = len(item["code"])
        if clen in by_len:
            by_len[clen].append(item)
    return code_to_name, items, by_len


CODE_TO_NAME, CHART, CHART_BY_LEN = build_chart()


def match_account(line: str, predicate=None):
    """Return the longest chart-item whose normalized form is a prefix of
    normalize(line).  The returned `rest` is the portion of the original line
    AFTER the matched name."""
    norm_line = normalize(line)
    candidates = CHART if predicate is None else CHART
    for item in candidates:
        if predicate and not predicate(item):
            continue
        if not item["norm"]:
            continue
        if norm_line.startswith(item["norm"]):
            rest = _walk_match(line, item["norm"])
            if rest is not None:
                return {
                    "code": item["code"],
                    "name": item["name"],
                    "rest": rest,
                }
    return None


AMOUNT_PATTERNS = [
    re.compile(r"合共\s*([\d,]+)\s*千\s*元"),
    re.compile(r"編列\s*([\d,]+)\s*千\s*元"),
    re.compile(r"([\d,]+)\s*千\s*元"),
]


def extract_amount(text: str):
    if not text:
        return ""
    for pat in AMOUNT_PATTERNS:
        m = pat.search(text)
        if m:
            try:
                return int(m.group(1).replace(",", "")) * 1000
            except ValueError:
                continue
    return ""


# ----- PDF extraction mirroring pdf.js y-sorted join -----

SECTION_MARKER = "基金用途明細表說明"
DATE_LINE_REGEX = re.compile(r"^中華民國\s*\d+\s*年度$")


def _render_page(page, y_tolerance: float) -> str:
    """Group words into lines by y-position (mirrors pdf.js y-sorted join)."""
    words = sorted(page.extract_words(), key=lambda w: (w["top"], w["x0"]))
    groups = []
    for w in words:
        if groups and abs(groups[-1]["y_ref"] - w["top"]) <= y_tolerance:
            groups[-1]["words"].append(w)
        else:
            groups.append({"y_ref": w["top"], "words": [w]})
    rendered = []
    for grp in groups:
        grp["words"].sort(key=lambda w: w["x0"])
        rendered.append("".join(w["text"] for w in grp["words"]))
    return "\n".join(rendered)


def _fund_of_page(text: str, section: str) -> str | None:
    """頁首「基金用途明細表說明」上一行即基金名稱；找不到回 None。"""
    head = [ln.replace(" ", "") for ln in text.split("\n")[:6]]
    if section in head:
        idx = head.index(section)
        if idx >= 1 and head[idx - 1]:
            return head[idx - 1]
    return None


def extract_pdf_segments(
    pdf_path: Path,
    y_tolerance: float = 3.0,
    section: str = SECTION_MARKER,
    page_range: tuple[int, int] | None = None,
) -> list[tuple[str, str]]:
    """回傳 [(基金名稱, 頁面文字), ...]，依頁首基金名稱把連續頁分段。

    - page_range 指定時只取該範圍；否則自動抓頁首含 section 標記的頁。
    - 頁面沒有基金頁首（例如 page_range 選到的頁）時沿用前一頁的基金；
      整份都無法判定時基金名為空字串。
    """
    with pdfplumber.open(pdf_path) as pdf:
        page_texts = [p.extract_text() or "" for p in pdf.pages]

        def _in_section(t: str) -> bool:
            head = t.split("\n")[:6]
            return any(ln.replace(" ", "") == section for ln in head)

        if page_range:  # 使用者指定 (起頁, 迄頁)，1-based 含迄頁，優先於自動偵測
            lo, hi = page_range
            idxs = list(range(max(lo, 1) - 1, min(hi, len(pdf.pages))))
        else:
            idxs = [i for i, t in enumerate(page_texts) if _in_section(t)] if section else []
            if not idxs:
                idxs = list(range(len(pdf.pages)))

        segments: list[tuple[str, list[str]]] = []
        current_fund = ""
        for i in idxs:
            fund = _fund_of_page(page_texts[i], section) or current_fund
            current_fund = fund
            text = _render_page(pdf.pages[i], y_tolerance)
            if segments and segments[-1][0] == fund:
                segments[-1][1].append(text)
            else:
                segments.append((fund, [text]))
    return [(fund, "\n".join(texts)) for fund, texts in segments]


def extract_pdf_text(
    pdf_path: Path,
    y_tolerance: float = 3.0,
    section: str = SECTION_MARKER,
    page_range: tuple[int, int] | None = None,
) -> str:
    """向後相容：只回傳文字（不含基金分段資訊）。"""
    segs = extract_pdf_segments(pdf_path, y_tolerance, section, page_range)
    return "\n".join(text for _, text in segs)


# ----- Cleanup pipeline (mirror of the three JS buttons) -----

HEADING_REGEX = re.compile(
    r"(?:(?<=^)|(?<=\s))"
    r"("
    r"[壹貳參肆伍陸柒捌玖拾]、|"
    r"[一二三四五六七八九十]、|"
    r"[^\s－，。：、]+(?:、[^\s－，。：、]+)*－|"
    r"\d+\.(?!\d)"
    r")"
)

PAREN_PREFIX_REGEX = re.compile(r"(\(\d{1,2}\)|\([一二三四五六七八九十]+\))")


def _split_at_period_outside_parens(text: str) -> str:
    """Insert \n after `。` UNLESS we are inside unmatched parens (so we don't
    chop a parenthetical clause in half)."""
    out = []
    depth = 0
    for c in text:
        if c in "（(":
            depth += 1
        elif c in "）)" and depth > 0:
            depth -= 1
        out.append(c)
        if c == "。" and depth == 0:
            out.append("\n")
    return "".join(out)


def break_lines_by_headings(text: str) -> str:
    one_line = re.sub(r"\s+", " ", text.replace("\n", " "))
    one_line = HEADING_REGEX.sub(r"\n\1", one_line)
    one_line = re.sub(r"\n([壹貳參肆伍陸柒捌玖拾]、)", r"\n\n\1", one_line)
    one_line = _split_at_period_outside_parens(one_line)
    return one_line.strip()


def break_before_parenthesis(text: str) -> str:
    return PAREN_PREFIX_REGEX.sub(r"\n  \1", text)


def clean_internal_whitespace(text: str) -> str:
    out = []
    for line in text.split("\n"):
        m = re.match(r"^(\s*)", line)
        lead = m.group(1) if m else ""
        body = line[len(lead):]
        body = re.sub(r"\s+", "", body).strip()  # 移除行內空白(含 PDF 換行造成的「千 元」空格)
        out.append(lead + body)
    return "\n".join(out)


FOOTER_REGEX = re.compile(r"^\s*\d+(-\d+)?\s*$")
DATE_HEADER_REGEX = re.compile(r"^中華民國\s*\d+\s*年度$")
PURE_CJK_REGEX = re.compile(r"^[一-鿿]{2,10}$")


def strip_page_chrome(text: str) -> str:
    lines = text.split("\n")
    # ponytail: 機關名/基金名等頁首行 = 短純中文行且整份重複出現 ≥3 次，動態偵測不寫死
    counts: dict[str, int] = {}
    for line in lines:
        t = line.strip().replace(" ", "")
        if PURE_CJK_REGEX.match(t):
            counts[t] = counts.get(t, 0) + 1
    out = []
    for line in lines:
        t = line.strip()
        if not t:
            continue
        compact = t.replace(" ", "")
        if FOOTER_REGEX.match(t):
            continue
        if compact == SECTION_MARKER or DATE_HEADER_REGEX.match(compact):
            continue
        if PURE_CJK_REGEX.match(compact) and counts.get(compact, 0) >= 3:
            continue
        out.append(line)
    return "\n".join(out)


def auto_clean(text: str) -> str:
    text = strip_page_chrome(text)
    text = break_lines_by_headings(text)
    text = break_before_parenthesis(text)
    text = clean_internal_whitespace(text)
    return text


# ----- Parser -----

PLAN_L1_REGEX = re.compile(r"^[壹貳參肆伍陸柒捌玖拾]、")
PLAN_L2_REGEX = re.compile(r"^[一二三四五六七八九十]、")
# 中文數字括號且整行不含金額 → 視為子計畫標題（第三層計畫），如「(一)精進豬隻保險業務計畫」
PLAN_L3_REGEX = re.compile(r"^[（(][一二三四五六七八九十]+[）)]")
NUMBER_PREFIX_REGEX = re.compile(r"^\s*(\d+)\.\s*")
PAREN_PREFIX_LINE_REGEX = re.compile(r"^\s*\((\d+|[一二三四五六七八九十]+)\)\s*")
CALC_NOTE_REGEX = re.compile(r"[=＝+＋\[\]％%]")


def budget_type_of(s: str):
    """預算別：同一計畫下的預算來源區分（原年度預算 vs 災後復原特別預算），非子計畫也非科目。"""
    t = s.strip()
    if t == "原年度預算":
        return "原年度預算"
    if "特別預算" in t or "災後復原重建" in t:
        return "特別預算"
    return None


def parse(text: str, fund: str = ""):
    rows = []
    unmatched = []
    current_plan_l1 = ""
    current_plan_l2 = ""
    current_plan_l3 = ""
    current_budget = ""
    current_l1 = ""
    current_l2 = ""
    current_l3 = ""

    for raw in text.split("\n"):
        line = raw.strip()
        if not line:
            continue

        if PLAN_L1_REGEX.match(line):
            current_plan_l1 = PLAN_L1_REGEX.sub("", line).strip()
            current_plan_l2 = current_plan_l3 = current_budget = ""
            current_l1 = current_l2 = current_l3 = ""
            continue
        if PLAN_L2_REGEX.match(line):
            current_plan_l2 = PLAN_L2_REGEX.sub("", line).strip()
            current_plan_l3 = current_budget = ""
            current_l1 = current_l2 = current_l3 = ""
            continue

        # 預算別標題（(一)原年度預算 / (二)因應…災後復原重建，或 1./2. 形式）：
        # 取標號後文字判定，整行不含金額才視為純標題
        if "千元" not in line:
            _mcn = PLAN_L3_REGEX.match(line)
            _mnum = NUMBER_PREFIX_REGEX.match(line)
            _htext = None
            if _mcn:
                _htext = line[_mcn.end():].strip()
            elif _mnum:
                _htext = line[_mnum.end():].strip()
            if _htext is not None:
                _bt = budget_type_of(_htext)
                if _bt:
                    current_budget = _bt
                    current_l1 = current_l2 = current_l3 = ""
                    continue

        if PLAN_L3_REGEX.match(line) and "千元" not in line:
            current_plan_l3 = PLAN_L3_REGEX.sub("", line).strip()
            current_l1 = current_l2 = current_l3 = ""
            continue

        # 沒有「編列」、又帶算式符號（= + [] %）的行 → 貸款餘額/利息計算說明，
        # 併回上一列說明，不當新科目也不抓金額
        if rows and "編列" not in line and CALC_NOTE_REGEX.search(line) and not match_account(
            PAREN_PREFIX_LINE_REGEX.sub("", NUMBER_PREFIX_REGEX.sub("", line))
        ):
            # 句首若是可比對到的科目(如「1.擴充改良房屋建築及設備…〈計算式x3%〉」)則不吞，讓它獨立成列
            prev = rows[-1]
            prev["description"] += "\n" + line
            prev["raw"] += "\n" + line
            continue

        work = line
        prefix_kind = "none"
        m = NUMBER_PREFIX_REGEX.match(line)
        if m:
            work = line[m.end():]
            prefix_kind = "number"
        else:
            m = PAREN_PREFIX_LINE_REGEX.match(line)
            if m:
                work = line[m.end():]
                prefix_kind = "paren"

        if prefix_kind == "none":
            level_pref = [1, 2, 3]
        elif prefix_kind == "number":
            level_pref = [2, 3]
        else:
            level_pref = [3]

        matched = None
        for lvl in level_pref:
            matched = match_account(work, lambda it, L=lvl: len(it["code"]) == L)
            if matched:
                break

        if not matched and prefix_kind == "number":
            norm_work = normalize(work)
            for lvl in level_pref:
                for item in CHART_BY_LEN.get(lvl, []):
                    if len(item["norm"]) < 5:
                        continue
                    if item["norm"] in norm_work:
                        matched = {"code": item["code"], "name": item["name"], "rest": ""}
                        break
                if matched:
                    break

        row_l1, row_l2, row_l3 = current_l1, current_l2, current_l3
        description = work

        if matched:
            clen = len(matched["code"])
            if clen == 1:
                current_l1 = matched["name"]
                current_l2 = current_l3 = ""
                row_l1 = matched["name"]
                row_l2 = ""
                row_l3 = ""
                rest = matched["rest"]
                if rest.startswith("－"):
                    rest = rest[1:]
                    l2m = match_account(
                        rest,
                        lambda it: len(it["code"]) == 2
                        and it["code"].startswith(matched["code"]),
                    )
                    if l2m:
                        current_l2 = l2m["name"]
                        row_l2 = l2m["name"]
                        rest = l2m["rest"]
                        if rest.startswith("－"):
                            rest = rest[1:]
                    else:
                        l3m = match_account(
                            rest,
                            lambda it: len(it["code"]) == 3
                            and it["code"].startswith(matched["code"]),
                        )
                        if l3m:
                            l2code = l3m["code"][:2]
                            current_l2 = CODE_TO_NAME.get(l2code, "")
                            current_l3 = l3m["name"]
                            row_l2 = current_l2
                            row_l3 = l3m["name"]
                            rest = l3m["rest"]
                            if rest.startswith("－"):
                                rest = rest[1:]
                description = rest.strip()
            elif clen == 2:
                l1code = matched["code"][:1]
                current_l1 = CODE_TO_NAME.get(l1code, current_l1)
                current_l2 = matched["name"]
                current_l3 = ""
                row_l1 = current_l1
                row_l2 = matched["name"]
                row_l3 = ""
                rest = matched["rest"]
                if rest.startswith("－"):
                    rest = rest[1:]
                description = rest.strip()
            else:
                l1code = matched["code"][:1]
                l2code = matched["code"][:2]
                current_l1 = CODE_TO_NAME.get(l1code, current_l1)
                current_l2 = CODE_TO_NAME.get(l2code, current_l2)
                current_l3 = matched["name"]
                row_l1 = current_l1
                row_l2 = current_l2
                row_l3 = matched["name"]
                rest = matched["rest"]
                if rest.startswith("－"):
                    rest = rest[1:]
                description = rest.strip()
        elif prefix_kind == "none" and rows:
            # 無科目、無標號 → 視為上一列說明的接續行，併入前列
            prev = rows[-1]
            prev["description"] += "\n" + line
            prev["raw"] += "\n" + line
            if not prev["amount"]:
                prev["amount"] = extract_amount(line)
            continue
        else:
            row_l1 = current_l1
            row_l2 = current_l2
            row_l3 = ""
            current_l3 = ""
            description = work.strip()
            if prefix_kind == "none":
                unmatched.append(line)

        amount = extract_amount(line)
        plan = " - ".join(p for p in (current_plan_l1, current_plan_l2, current_plan_l3) if p) or "未分類計畫"

        rows.append(
            {
                "fund": fund,
                "plan": plan,
                "budget": current_budget,
                "l1": row_l1,
                "l2": row_l2,
                "l3": row_l3,
                "amount": amount,
                "description": description or line,
                "raw": line,
            }
        )

    return rows, unmatched


def parse_pdf(
    pdf_path: Path,
    y_tolerance: float = 3.0,
    section: str = SECTION_MARKER,
    page_range: tuple[int, int] | None = None,
):
    """PDF → (rows, unmatched, cleaned_text)，每列標上所屬基金。"""
    all_rows, all_unmatched, cleaned_parts = [], [], []
    for fund_name, text in extract_pdf_segments(pdf_path, y_tolerance, section, page_range):
        cleaned = auto_clean(text)
        cleaned_parts.append(cleaned)
        rows, unmatched = parse(cleaned, fund=fund_name)
        all_rows.extend(rows)
        all_unmatched.extend(unmatched)
    return all_rows, all_unmatched, "\n".join(cleaned_parts)


# ----- xlsx output -----

def write_xlsx(rows, unmatched, out_path: Path, cleaned_text: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "解析結果"

    headers = [
        "基金",
        "計畫名稱",
        "預算別",
        "一級科目",
        "二級科目",
        "三級科目",
        "金額 (元)",
        "編列說明",
        "原始行",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", start_color="E9ECEF")
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, name="Arial")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append([r.get("fund", ""), r["plan"], r.get("budget", ""), r["l1"], r["l2"], r["l3"], r["amount"] or "", r["description"], r["raw"]])

    widths = [18, 24, 12, 16, 20, 18, 16, 80, 80]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=7)
        cell.number_format = "#,##0;(#,##0);-"
        cell.alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=8).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=row, column=9).alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"

    # Summary sheet — single pass
    ws2 = wb.create_sheet("摘要")
    ws2["A1"] = "項目"
    ws2["B1"] = "數值"
    ws2["A1"].font = ws2["B1"].font = Font(bold=True, name="Arial")
    with_amount = l1_only = l2 = l3 = 0
    for r in rows:
        if r["amount"]:
            with_amount += 1
        if r["l3"]:
            l3 += 1
        elif r["l2"]:
            l2 += 1
        elif r["l1"]:
            l1_only += 1
    ws2.append(["總列數", len(rows)])
    ws2.append(["含金額列", with_amount])
    ws2.append(["科目層級 1 行數", l1_only])
    ws2.append(["科目層級 2 行數", l2])
    ws2.append(["科目層級 3 行數", l3])
    ws2.append(["未匹配至科目行 (prefix=none)", len(unmatched)])
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 14

    if unmatched:
        ws3 = wb.create_sheet("未匹配行")
        ws3["A1"] = "行內容"
        ws3["A1"].font = Font(bold=True, name="Arial")
        for u in unmatched:
            ws3.append([u])
        ws3.column_dimensions["A"].width = 100
        for row in range(2, ws3.max_row + 1):
            ws3.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    ws4 = wb.create_sheet("清理後文字")
    ws4["A1"] = "行號"
    ws4["B1"] = "內容"
    ws4["A1"].font = ws4["B1"].font = Font(bold=True, name="Arial")
    for i, line in enumerate(cleaned_text.split("\n"), start=1):
        ws4.append([i, line])
    ws4.column_dimensions["A"].width = 8
    ws4.column_dimensions["B"].width = 120
    for row in range(2, ws4.max_row + 1):
        ws4.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws4.freeze_panes = "A2"

    wb.save(out_path)


def main():
    pdf_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
        "農業基金用途明細114年度預算案 (1).pdf"
    )
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(
        "農業基金用途明細114年度預算案_解析結果.xlsx"
    )

    rows, unmatched, cleaned = parse_pdf(pdf_path)
    write_xlsx(rows, unmatched, out_path, cleaned)

    funds = {}
    for r in rows:
        funds[r.get("fund", "")] = funds.get(r.get("fund", ""), 0) + 1
    print(f"Wrote {out_path}")
    print(f"rows={len(rows)}  unmatched(prefix=none)={len(unmatched)}")
    print("基金分布:", funds)


if __name__ == "__main__":
    main()
